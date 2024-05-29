import sqlite3

import openpyxl

wb = openpyxl.load_workbook("table.russia.xlsx")
input_ws = wb["Лист1"]
cities_tuple: tuple[str] = tuple(col[0].lower() for col in input_ws.iter_cols(min_row=1, max_row=1, min_col=2,
                                                                              max_col=input_ws.max_column, values_only=True))
connection = sqlite3.connect("cities.db")
cursor = connection.cursor()

connection.execute("CREATE TABLE IF NOT EXISTS cities (city TEXT PRIMARY KEY);")
for col_index, city in enumerate(cities_tuple, start=2):
    connection.execute(f"INSERT INTO cities VALUES ('{city}');")
    connection.execute(f"CREATE TABLE IF NOT EXISTS '{city}' (city TEXT PRIMARY KEY, distance INTEGER);")
    for row_index, distance in enumerate(input_ws.iter_rows(min_row=2, max_row=input_ws.max_row, min_col=col_index, max_col=col_index, values_only=True)):
        cursor.execute(f"INSERT INTO '{city}' (city, distance) VALUES (?, ?);",
                       (cities_tuple[row_index],  0 if isinstance(distance[0], str) else distance[0]))
connection.commit()
